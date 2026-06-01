"""
/listapromofim — Relatório de promoções ativas com análise de continuidade.
Versão otimizada com requisições paralelas.
"""

import time
import sys
from datetime import datetime, date, timezone
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from auth import get_valid_token

BASE = "https://api.mercadolibre.com"
USER_ID = 48980675
MAX_WORKERS = 10

_token = None


def _headers():
    global _token
    if not _token:
        _token = get_valid_token()
    return {"Authorization": f"Bearer {_token}"}


def _get(path, params=None, retries=3):
    for attempt in range(retries):
        try:
            r = requests.get(f"{BASE}{path}", headers=_headers(), params=params, timeout=15)
            if r.status_code == 429:
                time.sleep(2 ** (attempt + 1))
                continue
            r.raise_for_status()
            return r.json()
        except Exception:
            if attempt == retries - 1:
                return None
            time.sleep(1)


def _fetch_item_promos(item_id):
    data = _get(f"/seller-promotions/items/{item_id}", params={"app_version": "v2"})
    started, pending, price_matching = [], [], []
    if data:
        items = data if isinstance(data, list) else data.get("results", [])
        for p in items:
            s = p.get("status")
            if p.get("type") == "PRICE_MATCHING" and s in ("started", "candidate"):
                price_matching.append(p)
            if s == "started":
                started.append(p)
            elif s in ("pending", "candidate"):
                pending.append(p)
    return item_id, started, pending, price_matching


def run():
    t0 = time.time()
    print("\n=== /listapromofim — TSC Shop ===\n", flush=True)

    # --- Passo 1: Promoções do seller ---
    print("1/6 Promoções do seller...", flush=True)
    promos_started, promos_pending = {}, {}
    offset = 0
    while True:
        data = _get(f"/seller-promotions/users/{USER_ID}", params={
            "app_version": "v2", "limit": 50, "offset": offset
        })
        if not data:
            break
        results = data.get("results", [])
        if not results:
            break
        for p in results:
            pid = p.get("id")
            st = p.get("status")
            if st in ("started", "candidate"):
                promos_started[pid] = p
            elif st == "pending":
                promos_pending[pid] = p
        if len(results) < 50:
            break
        offset += 50
    print(f"     Ativas: {len(promos_started)} | Pendentes: {len(promos_pending)}", flush=True)

    # --- Passo 2: Todos os itens ---
    print("2/6 Listando anúncios...", flush=True)
    todos = []
    params = {"search_type": "scan", "limit": 100}
    while True:
        data = _get(f"/users/{USER_ID}/items/search", params=params)
        if not data:
            break
        ids = data.get("results", [])
        todos.extend(ids)
        scroll_id = data.get("scroll_id")
        if not scroll_id or not ids:
            break
        params = {"search_type": "scan", "scroll_id": scroll_id, "limit": 100}
    print(f"     {len(todos)} itens", flush=True)

    # --- Passo 2b: Itens vendidos nos últimos 30 dias ---
    print("2b/6 Itens vendidos nos últimos 30 dias...", flush=True)
    from datetime import timedelta
    data_30d = (date.today() - timedelta(days=30)).strftime("%Y-%m-%dT00:00:00.000-03:00")
    itens_vendidos_30d = set()
    offset_v = 0
    while True:
        data_v = _get(f"/orders/search", params={
            "seller": USER_ID,
            "order.date_created.from": data_30d,
            "order.status": "paid",
            "limit": 50,
            "offset": offset_v,
        })
        if not data_v:
            break
        orders = data_v.get("results", [])
        for o in orders:
            for oi in o.get("order_items", []):
                iid = oi.get("item", {}).get("id")
                if iid:
                    itens_vendidos_30d.add(iid)
        total_v = data_v.get("paging", {}).get("total", 0)
        offset_v += 50
        if offset_v >= total_v:
            break
    print(f"     {len(itens_vendidos_30d)} itens distintos vendidos", flush=True)

    # --- Passo 3: Promos por item (PARALELO) ---
    print(f"3/6 Promoções por item ({MAX_WORKERS} threads)...", flush=True)
    item_promos_started = defaultdict(list)
    item_promos_pending = defaultdict(list)
    item_price_matching = {}
    done = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_fetch_item_promos, iid): iid for iid in todos}
        for f in as_completed(futures):
            done += 1
            if done % 100 == 0:
                print(f"     {done}/{len(todos)}...", flush=True)
            item_id, started, pending, price_matching = f.result()
            for p in started:
                item_promos_started[item_id].append(p)
            for p in pending:
                item_promos_pending[item_id].append(p)
            if price_matching:
                item_price_matching[item_id] = price_matching[0]
    print(f"     {len(item_promos_started)} itens com promo ativa", flush=True)

    # Merge seller-level e injetar finish_date nos promos por item
    all_seller_promos = {**promos_started, **promos_pending}
    for item_id, promos in item_promos_started.items():
        for p in promos:
            pid = p.get("id")
            if pid and pid in all_seller_promos:
                seller_p = all_seller_promos[pid]
                if not p.get("finish_date") and seller_p.get("finish_date"):
                    p["finish_date"] = seller_p["finish_date"]
                if not p.get("end_time") and seller_p.get("deadline_date"):
                    p["end_time"] = seller_p["deadline_date"]

    for pid, promo in promos_started.items():
        for i in promo.get("items", []):
            iid = i.get("item_id") or i.get("id")
            if iid and promo not in item_promos_started[iid]:
                item_promos_started[iid].append(promo)

    # --- Passo 4: Títulos (batch de 20) — todos os itens ---
    itens_com_promo = set(item_promos_started.keys())
    itens_sem_promo = [iid for iid in todos if iid not in itens_com_promo]
    todos_ids_para_titulo = list(itens_com_promo) + itens_sem_promo
    print(f"4/6 Títulos ({len(todos_ids_para_titulo)} itens)...", flush=True)
    titulos = {}
    for i in range(0, len(todos_ids_para_titulo), 20):
        batch = todos_ids_para_titulo[i:i+20]
        data = _get("/items", params={"ids": ",".join(batch), "attributes": "id,title,price,catalog_listing,catalog_product_id"})
        if data:
            for item in (data if isinstance(data, list) else []):
                body = item.get("body", item)
                if body:
                    titulos[body.get("id")] = {
                        "title": body.get("title", ""),
                        "price": body.get("price", 0),
                        "catalog_listing": body.get("catalog_listing", False),
                        "catalog_product_id": body.get("catalog_product_id", ""),
                    }

    # --- Passo 5: Análise ---
    print("5/6 Analisando continuidade...", flush=True)
    rows = []
    for item_id, promos in item_promos_started.items():
        info = titulos.get(item_id, {})
        titulo = info.get("title", item_id)
        preco_orig = info.get("price", 0)

        datas_fim = []
        for p in promos:
            fd = p.get("finish_date") or p.get("end_time")
            if fd:
                try:
                    dt = datetime.fromisoformat(fd.replace("Z", "+00:00"))
                    if dt.tzinfo is None:
                        from datetime import timezone
                        dt = dt.replace(tzinfo=timezone.utc)
                    datas_fim.append(dt)
                except Exception:
                    pass

        earliest_finish = min(datas_fim) if datas_fim else None
        latest_finish = max(datas_fim) if datas_fim else None

        pendentes = item_promos_pending.get(item_id, [])
        has_continuity = bool(pendentes) or len(set(str(d) for d in datas_fim)) > 1 or any(
            p.get("status") == "candidate" for p in promos
        )

        promo_names = [p.get("name") or p.get("type", "?") for p in promos]
        tipos = [p.get("type", "?") for p in promos]
        preco_promo = min((p.get("new_price") or p.get("offer_price") or preco_orig for p in promos), default=preco_orig)
        desconto = round((1 - preco_promo / preco_orig) * 100, 1) if preco_orig else 0
        proximas = [p.get("name") or p.get("type", "?") for p in pendentes]

        item_info = titulos.get(item_id, {})
        rows.append({
            "item_id": item_id,
            "titulo": titulo,
            "promo_ativa": ", ".join(set(promo_names)),
            "tipo": ", ".join(set(tipos)),
            "preco_orig": preco_orig,
            "preco_promo": preco_promo,
            "desconto": desconto,
            "fim_promo": earliest_finish.strftime("%d/%m/%Y %H:%M") if earliest_finish else "-",
            "cobertura_ate": latest_finish.strftime("%d/%m/%Y %H:%M") if latest_finish else "-",
            "continuidade": "SIM" if has_continuity else "NÃO",
            "proximas": ", ".join(proximas) if proximas else "-",
            "_earliest": earliest_finish or datetime.max.replace(tzinfo=timezone.utc),
            "_catalog_listing": item_info.get("catalog_listing", False),
            "_catalog_product_id": item_info.get("catalog_product_id", ""),
        })

    # Deduplicar: por catalog_product_id, preferir o de catálogo
    from collections import defaultdict as _dd
    grupos = _dd(list)
    sem_catalogo_id = []
    for r in rows:
        if r["_catalog_product_id"]:
            grupos[r["_catalog_product_id"]].append(r)
        else:
            sem_catalogo_id.append(r)
    rows_dedup = sem_catalogo_id[:]
    for cpid, grupo in grupos.items():
        catalogo = [r for r in grupo if r["_catalog_listing"]]
        rows_dedup.append(catalogo[0] if catalogo else grupo[0])

    rows = rows_dedup
    rows.sort(key=lambda x: x["_earliest"].replace(tzinfo=None) if x["_earliest"].tzinfo else x["_earliest"])

    # --- Passo 6: Excel ---
    print("6/6 Gerando Excel...", flush=True)
    hoje = date.today().strftime("%Y%m%d")
    caminho = f"C:\\Users\\Henrique Dereste\\tsc\\promocoes_tsc_{hoje}.xlsx"

    wb = openpyxl.Workbook()
    h_fill = PatternFill("solid", fgColor="2F5496")
    h_font = Font(color="FFFFFF", bold=True)
    verde = PatternFill("solid", fgColor="C6EFCE")
    vermelho = PatternFill("solid", fgColor="FFC7CE")

    # Aba 1: Participando
    ws1 = wb.active
    ws1.title = "Participando"
    cab1 = ["Item ID", "Título", "Promo Ativa", "Tipo", "Preço Orig", "Preço Promo",
            "Desconto %", "Fim Promo", "Cobertura Até", "Continuidade?", "Próximas"]
    for col, txt in enumerate(cab1, 1):
        c = ws1.cell(1, col, txt)
        c.fill, c.font = h_fill, h_font
    for row in rows:
        ws1.append([row["item_id"], row["titulo"], row["promo_ativa"], row["tipo"],
                     row["preco_orig"], row["preco_promo"], row["desconto"],
                     row["fim_promo"], row["cobertura_ate"], row["continuidade"], row["proximas"]])
        ws1.cell(ws1.max_row, 10).fill = verde if row["continuidade"] == "SIM" else vermelho
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:K{ws1.max_row}"
    for col in ws1.columns:
        ws1.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(len(str(c.value or "")) for c in col) + 3, 60)

    # Aba 2: Sem Promoção
    ws_sem = wb.create_sheet("Sem Promoção (vendidos 30d)")
    h_gray = PatternFill("solid", fgColor="595959")
    cab_sem = ["Item ID", "Título", "Preço Atual"]
    for col, txt in enumerate(cab_sem, 1):
        c = ws_sem.cell(1, col, txt)
        c.fill, c.font = h_gray, h_font
    # Filtrar sem promoção: só itens vendidos nos últimos 30 dias
    itens_sem_promo = [iid for iid in itens_sem_promo if iid in itens_vendidos_30d]

    # Deduplicar sem promoção pelo mesmo critério
    grupos_sem = _dd(list)
    sem_cid = []
    for iid in itens_sem_promo:
        info = titulos.get(iid, {})
        cpid = info.get("catalog_product_id", "")
        entry = (iid, info.get("title", iid), info.get("price", ""), info.get("catalog_listing", False), cpid)
        if cpid:
            grupos_sem[cpid].append(entry)
        else:
            sem_cid.append(entry)
    itens_sem_dedup = sem_cid[:]
    for cpid, grupo in grupos_sem.items():
        catalogo = [e for e in grupo if e[3]]
        itens_sem_dedup.append(catalogo[0] if catalogo else grupo[0])

    rows_sem = [[e[0], e[1], e[2]] for e in itens_sem_dedup]
    rows_sem.sort(key=lambda x: x[1])
    for r in rows_sem:
        ws_sem.append(r)
    ws_sem.freeze_panes = "A2"
    ws_sem.auto_filter.ref = f"A1:C{ws_sem.max_row}"
    for col in ws_sem.columns:
        ws_sem.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(len(str(c.value or "")) for c in col) + 3, 60)

    # Aba 3: Resumo
    ws2 = wb.create_sheet("Resumo")
    ws2.append(["Fim Cobertura", "Total Itens", "Com Continuidade", "SEM Continuidade"])
    for c in ws2[1]:
        c.fill, c.font = h_fill, h_font
    for d in sorted(set(r["cobertura_ate"] for r in rows if r["cobertura_ate"] != "-")):
        grupo = [r for r in rows if r["cobertura_ate"] == d]
        sim = sum(1 for r in grupo if r["continuidade"] == "SIM")
        nao = len(grupo) - sim
        ws2.append([d, len(grupo), sim, nao])
        if nao > 0:
            ws2.cell(ws2.max_row, 4).fill = vermelho

    # Aba 3: SEM Continuidade
    ws3 = wb.create_sheet("SEM Continuidade - ALERTA")
    h_red = PatternFill("solid", fgColor="C00000")
    cab3 = ["Item ID", "Título", "Promo(s) Ativa(s)", "Fim Cobertura", "Preço Orig", "Preço Promo", "Desconto %"]
    for col, txt in enumerate(cab3, 1):
        c = ws3.cell(1, col, txt)
        c.fill, c.font = h_red, h_font
    sem = [r for r in rows if r["continuidade"] == "NÃO"]
    for r in sem:
        ws3.append([r["item_id"], r["titulo"], r["promo_ativa"], r["cobertura_ate"],
                     r["preco_orig"], r["preco_promo"], r["desconto"]])
    ws3.auto_filter.ref = f"A1:G{ws3.max_row}"
    for col in ws3.columns:
        ws3.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(len(str(c.value or "")) for c in col) + 3, 60)

    # Aba 5: Desconto na Tarifa (PRICE_MATCHING)
    ws_pm = wb.create_sheet("Desconto na Tarifa")
    h_blue = PatternFill("solid", fgColor="1F4E79")
    cab_pm = ["Item ID", "Título", "Preço Orig", "Preço Promo", "% Seller", "% ML", "Participando?"]
    for col, txt in enumerate(cab_pm, 1):
        c = ws_pm.cell(1, col, txt)
        c.fill, c.font = h_blue, h_font

    azul_claro = PatternFill("solid", fgColor="BDD7EE")

    # Buscar títulos de todos os itens (já estão em titulos{})
    todos_pm = []
    for iid in todos:
        info = titulos.get(iid, {})
        pm = item_price_matching.get(iid)
        todos_pm.append({
            "item_id": iid,
            "titulo": info.get("title", iid),
            "preco_orig": info.get("price", ""),
            "preco_promo": pm.get("price", "-") if pm else "-",
            "seller_pct": pm.get("seller_percentage", "-") if pm else "-",
            "meli_pct": pm.get("meli_percentage", "-") if pm else "-",
            "participando": "SIM" if pm else "NÃO",
        })

    todos_pm.sort(key=lambda x: (x["participando"] == "NÃO", x["titulo"]))
    for r in todos_pm:
        ws_pm.append([r["item_id"], r["titulo"], r["preco_orig"], r["preco_promo"],
                      r["seller_pct"], r["meli_pct"], r["participando"]])
        ws_pm.cell(ws_pm.max_row, 7).fill = azul_claro if r["participando"] == "SIM" else vermelho

    ws_pm.freeze_panes = "A2"
    ws_pm.auto_filter.ref = f"A1:G{ws_pm.max_row}"
    for col in ws_pm.columns:
        ws_pm.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(len(str(c.value or "")) for c in col) + 3, 60)

    wb.save(caminho)

    elapsed = time.time() - t0
    print(f"\n{'='*45}")
    print(f"  Total itens em promoção : {len(rows)}")
    print(f"  COM continuidade        : {sum(1 for r in rows if r['continuidade']=='SIM')}")
    print(f"  SEM continuidade (alerta): {len(sem)}")
    print(f"  Sem promoção (vendidos 30d): {len(itens_sem_promo)}")
    print(f"  Tempo: {elapsed:.0f}s")
    print(f"  Excel: {caminho}")
    print(f"{'='*45}\n")


if __name__ == "__main__":
    run()
